"""
pokemon_pipeline.py

One-stop script that:
  1. Fills the "Icon" column (D) on the 'Teams' sheet of Pokemon.xlsx with a
     sprite image for every Pokemon name in column C (from add_pokemon_icons.py).
  2. Reads Pokemon.xlsx and rebuilds pokemon_ui.html directly by injecting
     the parsed data into pokemon_ui_template.html - no intermediate JSON
     file is written; this is a static site with no server, so there's
     nothing that needs to read a standalone data.json at runtime.

Keep these files in the same folder:
  Pokemon.xlsx              <- your spreadsheet (edit this)
  pokemon_ui_template.html  <- page skeleton (don't edit unless changing layout)
  pokemon_pipeline.py       <- this script
  pokemon_ui.html           <- generated, open this in a browser

USAGE
-----
    python pokemon_pipeline.py

That's the only command you need. Every time you edit Pokemon.xlsx (add/change
Pokemon names, teams, etc.), just re-run this script - it will refresh the
icons in the workbook, then regenerate pokemon_ui.html.

Requires: pip install openpyxl requests pillow
"""

import io
import json
import re
import time
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import requests
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image as PILImage

# ---------------------------------------------------------------------------
# Shared config
# ---------------------------------------------------------------------------

SOURCE_FILE = "Pokemon.xlsx"
XLSX_OUTPUT_FILE = "Pokemon.xlsx"          # overwrite in place after adding icons
TEMPLATE_FILE = "pokemon_ui_template.html"
HTML_OUTPUT_FILE = "pokemon_ui.html"

SHEET_NAME = "Teams"
NAME_COL = 3                          # column C - Pokemon
ICON_COL = 4                          # column D - Icon
FIRST_DATA_ROW = 2


# ===========================================================================
# STEP 1 - add sprite icons into the Excel workbook (Teams!D)
# ===========================================================================

ICON_PX = 40                           # rendered icon size in pixels
CELL_PADDING_PX = 6                    # breathing room around the icon
CACHE_DIR = Path(".pokemon_icon_cache")

SPRITE_URL = "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/{id}.png"
POKEAPI_URL = "https://pokeapi.co/api/v2/pokemon/{name}"

DEFAULT_COL_WIDTH = 8.43
DEFAULT_ROW_HEIGHT = 15.0

# Sheets/ranges that need an in-cell dropdown letting you pick any Pokemon
# name from the Teams sheet's Pokemon column (C). openpyxl can't preserve
# Excel's "extended" cross-sheet list validation on load/save, so this
# recreates a plain (non-extended) List validation with the same effect
# every time the script runs.
DROPDOWN_TARGETS = {
    "Specialising - Type": ["D2:I100"],
    "Specialising - Gen": ["D2:I100"],
    "Specialising - Gen_Type": ["B2:J100"],
}
DROPDOWN_SOURCE = "=Teams!$C$2:$C$1000"

# Forms/special-name overrides where the display name in the sheet doesn't
# match the PokeAPI slug exactly. Used for fetching icon sprites (PokeAPI).
ICON_NAME_OVERRIDES = {
    "aegislash": "aegislash-shield",
    "mimikyu": "mimikyu-disguised",
    "palafin": "palafin-zero",
}

# Overrides used for the pokemondb.net sprite URLs embedded in the UI JSON.
UI_NAME_OVERRIDES = {
    "aegislash": "aegislash-shield",
    "palafin": "palafin-zero",
    "tauros-paldea-aqua-breed": "tauros-paldean-aqua",
    "zoroark-hisui": "zoroark-hisuian",
    "ninetales-alola": "ninetales-alolan",
    "raichu-alola": "raichu-alolan",
}


def icon_slugify(name: str) -> str:
    n = name.strip().lower()
    if n in ICON_NAME_OVERRIDES:
        return ICON_NAME_OVERRIDES[n]
    n = re.sub(r"[.'’]", "", n)
    n = re.sub(r"[^a-z0-9]+", "-", n).strip("-")
    return n


def fetch_icon_bytes(name: str):
    """Download (and cache) a small PNG sprite for a Pokemon name. Returns None if not found."""
    CACHE_DIR.mkdir(exist_ok=True)
    slug = icon_slugify(name)
    cache_path = CACHE_DIR / f"{slug}.png"
    if cache_path.exists():
        return cache_path.read_bytes()

    for attempt in range(3):
        try:
            r = requests.get(POKEAPI_URL.format(name=slug), timeout=10)
            if r.status_code != 200:
                continue
            data = r.json()
            sprite_url = data["sprites"]["front_default"]
            if not sprite_url:
                sprite_url = SPRITE_URL.format(id=data["id"])
            img_resp = requests.get(sprite_url, timeout=10)
            if img_resp.status_code != 200:
                continue
            cache_path.write_bytes(img_resp.content)
            return img_resp.content
        except requests.RequestException:
            time.sleep(0.5)
            continue
    return None


def make_thumbnail(png_bytes: bytes, size: int) -> io.BytesIO:
    im = PILImage.open(io.BytesIO(png_bytes)).convert("RGBA")

    if im.width < size and im.height < size:
        canvas = PILImage.new("RGBA", (size, size), (0, 0, 0, 0))
        offset = ((size - im.width) // 2, (size - im.height) // 2)
        canvas.paste(im, offset, im)
        im = canvas
    else:
        im = im.resize((size, size), PILImage.Resampling.LANCZOS)

    buf = io.BytesIO()
    im.save(buf, format="PNG")
    buf.seek(0)
    return buf


def col_width_to_px(width_chars: float) -> float:
    return round(width_chars * 7 + 5)


def px_to_col_width(px: float) -> float:
    return (px - 5) / 7


def row_height_to_px(height_pts: float) -> float:
    return height_pts * 96 / 72


def px_to_row_height(px: float) -> float:
    return px * 72 / 96


def ensure_cell_fits_icon(ws, col_idx: int, row_idx: int, icon_px: int, padding_px: int):
    """Grows the column width / row height (never shrinks them) so the icon
    isn't clipped or cramped."""
    needed_px = icon_px + padding_px * 2

    col_letter = get_column_letter(col_idx + 1)
    col_dim = ws.column_dimensions[col_letter]
    current_width_chars = col_dim.width if col_dim.width else DEFAULT_COL_WIDTH
    needed_width_chars = px_to_col_width(needed_px)
    if needed_width_chars > current_width_chars:
        col_dim.width = needed_width_chars

    row_dim = ws.row_dimensions[row_idx + 1]
    current_height_pts = row_dim.height if row_dim.height else DEFAULT_ROW_HEIGHT
    needed_height_pts = px_to_row_height(needed_px)
    if needed_height_pts > current_height_pts:
        row_dim.height = needed_height_pts


def centered_anchor(ws, col_idx: int, row_idx: int, icon_px: int) -> OneCellAnchor:
    """Build a OneCellAnchor that centers an icon_px x icon_px image inside
    the given cell (0-indexed col_idx/row_idx)."""
    col_letter = get_column_letter(col_idx + 1)
    col_dim = ws.column_dimensions.get(col_letter)
    width_chars = col_dim.width if (col_dim and col_dim.width) else DEFAULT_COL_WIDTH

    row_dim = ws.row_dimensions.get(row_idx + 1)
    height_pts = row_dim.height if (row_dim and row_dim.height) else DEFAULT_ROW_HEIGHT

    cell_w_px = col_width_to_px(width_chars)
    cell_h_px = row_height_to_px(height_pts)

    off_x_px = max((cell_w_px - icon_px) / 2, 0)
    off_y_px = max((cell_h_px - icon_px) / 2, 0)

    marker = AnchorMarker(
        col=col_idx, colOff=pixels_to_EMU(off_x_px),
        row=row_idx, rowOff=pixels_to_EMU(off_y_px),
    )
    size = XDRPositiveSize2D(pixels_to_EMU(icon_px), pixels_to_EMU(icon_px))
    return OneCellAnchor(_from=marker, ext=size)


def restore_pokemon_dropdowns(wb):
    """Re-adds the Pokemon-name dropdown (list validation, source = Teams
    column C) to the Specialising sheets - openpyxl strips Excel's extended
    cross-sheet validation format on save, so this recreates a standard
    version of it every run."""
    for sheet_name, ranges in DROPDOWN_TARGETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        target_ws = wb[sheet_name]
        dv = DataValidation(type="list", formula1=DROPDOWN_SOURCE, allow_blank=True)
        target_ws.add_data_validation(dv)
        for rng in ranges:
            dv.add(rng)


def add_icons_to_workbook():
    """Step 1: fill Teams!D with sprite images, overwriting Pokemon.xlsx."""
    wb = openpyxl.load_workbook(SOURCE_FILE)
    ws = wb[SHEET_NAME]

    if ws.cell(1, ICON_COL).value != "Icon":
        raise SystemExit(
            f"Expected '{get_column_letter(ICON_COL)}1' to be 'Icon' but found "
            f"{ws.cell(1, ICON_COL).value!r}. Check ICON_COL / NAME_COL settings "
            f"match your sheet layout."
        )

    # remove any icon images we previously placed, so reruns stay in sync
    icon_col_letter = get_column_letter(ICON_COL)
    keep_images = []
    for img in ws._images:
        anchor_cell = img.anchor._from
        col_letter = get_column_letter(anchor_cell.col + 1)
        if col_letter == icon_col_letter:
            continue
        keep_images.append(img)
    ws._images = keep_images

    last_row = ws.max_row
    placed = 0
    misses = []
    for row in range(FIRST_DATA_ROW, last_row + 1):
        name = ws.cell(row, NAME_COL).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        png_bytes = fetch_icon_bytes(name)
        if png_bytes is None:
            misses.append((row, name))
            continue
        ensure_cell_fits_icon(ws, ICON_COL - 1, row - 1, ICON_PX, CELL_PADDING_PX)
        thumb = make_thumbnail(png_bytes, ICON_PX)
        img = XLImage(thumb)
        img.width = ICON_PX
        img.height = ICON_PX
        img.anchor = centered_anchor(ws, ICON_COL - 1, row - 1, ICON_PX)
        ws.add_image(img)
        placed += 1
        time.sleep(0.05)  # be polite to the API

    restore_pokemon_dropdowns(wb)

    wb.save(XLSX_OUTPUT_FILE)
    print(f"[2/2] Saved {XLSX_OUTPUT_FILE} - placed {placed} icon(s)")
    if misses:
        print("      Could not find an icon for these rows (check spelling / add to ICON_NAME_OVERRIDES):")
        for row, name in misses:
            print(f"        row {row}: {name!r}")


# ===========================================================================
# STEP 2 - build pokemon_ui.html directly from the workbook (no JSON file)
# ===========================================================================

def ui_slugify(name: str) -> str:
    n = str(name).strip().lower()
    if n in UI_NAME_OVERRIDES:
        return UI_NAME_OVERRIDES[n]
    n = re.sub(r"[.'’]", "", n)
    n = re.sub(r"[^a-z0-9]+", "-", n).strip("-")
    return n


def sprite_url(name: str) -> str:
    slug = ui_slugify(name)
    return f"https://img.pokemondb.net/sprites/home/normal/{slug}.png"


def forward_fill_merged(ws, col_idx):
    """Return {row: value} for a column, filling merged-cell gaps by
    propagating the top value of each merged range down through its rows."""
    values = {}
    for row in range(1, ws.max_row + 1):
        values[row] = ws.cell(row, col_idx).value
    for mrange in ws.merged_cells.ranges:
        if mrange.min_col == col_idx and mrange.max_col == col_idx:
            top_val = ws.cell(mrange.min_row, col_idx).value
            for r in range(mrange.min_row, mrange.max_row + 1):
                values[r] = top_val
    return values


def team_blocks(ws):
    """Team boundaries come from the merged cells in column B (Team name),
    not from matching adjacent values."""
    ranges = [r for r in ws.merged_cells.ranges if r.min_col == 2 and r.max_col == 2]
    ranges.sort(key=lambda r: r.min_row)
    blocks = []
    seen_rows = set()
    for r in ranges:
        blocks.append((r.min_row, r.max_row))
        seen_rows.update(range(r.min_row, r.max_row + 1))
    row = 2
    while row <= ws.max_row:
        if row not in seen_rows and ws.cell(row, 2).value:
            blocks.append((row, row))
            seen_rows.add(row)
        row += 1
    blocks.sort()
    return blocks


RED_RGB = "FFFF0000"
BLUE_RGB = "FF0070C0"


def font_rgb(cell):
    """Returns the cell's explicit RGB font color (e.g. 'FFFF0000') if set,
    or None if the font uses a theme/automatic color (i.e. default black)."""
    color = cell.font.color
    if not color:
        return None
    try:
        if color.type == "rgb" and isinstance(color.rgb, str):
            return color.rgb.upper()
    except AttributeError:
        pass
    return None


def export_teams(wb):
    ws = wb["Teams"]
    region_col = forward_fill_merged(ws, 1)
    teams = []
    for min_row, max_row in team_blocks(ws):
        team_name = ws.cell(min_row, 2).value
        region = region_col.get(min_row)
        team_is_red = font_rgb(ws.cell(min_row, 2)) == RED_RGB
        pokemons = []
        for row in range(min_row, max_row + 1):
            name = ws.cell(row, 3).value
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            types = [ws.cell(row, c).value for c in (5, 6)]
            types = [t for t in types if t]
            moves = []
            for c in (7, 8, 9, 10):
                cell = ws.cell(row, c)
                if cell.value:
                    moves.append({
                        "name": cell.value,
                        "underline": cell.font.underline is not None,
                    })
            gen = ws.cell(row, 11).value
            name_rgb = font_rgb(ws.cell(row, 3))
            if team_is_red or name_rgb == RED_RGB:
                star = "gold"
            elif name_rgb == BLUE_RGB:
                star = "silver"
            else:
                star = None
            pokemons.append({
                "pokemon": name,
                "sprite": sprite_url(name),
                "types": types,
                "moves": moves,
                "generation": str(gen).strip() if gen is not None else None,
                "star": star,
            })
        if not pokemons:
            continue
        teams.append({
            "region": region,
            "team": team_name,
            "pokemons": pokemons,
        })
    summary = compute_summary_stats(ws)
    return teams, summary


def compute_summary_stats(ws):
    """Mirrors the header stat formulas in Teams!N2:N4 (unique Pokemon count,
    raw team-name-cell count, unique team-name count) directly from columns
    B (Team) and C (Pokemon) - these are also formulas and subject to the
    same caching bug as everything else, so we recompute them instead of
    reading N2:N4."""
    pokemon_names = [
        str(ws.cell(row, 3).value).strip()
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 3).value not in (None, "")
    ]
    team_names = [
        str(ws.cell(row, 2).value).strip()
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 2).value not in (None, "")
    ]
    return {
        "pokemons_used": len(set(pokemon_names)),
        "teams_used": len(team_names),
        "games_played": len(set(team_names)),
    }


def _mon_ref(name):
    if not name:
        return None
    return {"name": name, "sprite": sprite_url(name)}


def build_roster_stats(teams):
    """Recompute the same numbers as the Excel formulas (COUNTIF/SUMPRODUCT
    over the Teams sheet) directly from the already-parsed roster, instead of
    trusting Excel's cached formula results.

    Why: openpyxl cannot preserve a formula's cached calculated value when it
    re-saves a workbook (e.g. the icon-adding step) - it only keeps the
    formula text. Any later read with data_only=True then sees None for
    every formula-driven cell (Mons/Used Count, Moves Count, the pivots)
    until the file is reopened and saved in real Excel. Since "Ace"/"Team"
    columns are plain typed values (not formulas) they're unaffected, but
    the count columns aren't - so we stop depending on them and compute the
    same figures ourselves from the raw Pokemon/type/generation/move data,
    which is immune to this bug no matter how many times the workbook gets
    resaved by openpyxl.
    """
    type_used_count = Counter()      # mirrors COUNTIF(Teams!E:F, type) - counts cell matches
    type_mons = defaultdict(set)     # unique Pokemon names per type
    gen_used_count = Counter()       # mirrors COUNTIF(Teams!K:K, gen)
    gen_mons = defaultdict(set)      # unique Pokemon names per generation
    move_used_count = Counter()      # mirrors COUNTIF(Teams!G:J, move name)

    for team in teams:
        for p in team["pokemons"]:
            name = p["pokemon"]
            for t in p["types"]:
                type_used_count[t] += 1
                type_mons[t].add(name)
            gen = p["generation"]
            if gen:
                gen_used_count[gen] += 1
                gen_mons[gen].add(name)
            for mv in p["moves"]:
                move_used_count[mv["name"]] += 1

    return {
        "type_used_count": type_used_count,
        "type_mons_count": {t: len(names) for t, names in type_mons.items()},
        "gen_used_count": gen_used_count,
        "gen_mons_count": {g: len(names) for g, names in gen_mons.items()},
        "move_used_count": move_used_count,
    }


def export_specialising_type(wb, stats):
    ws = wb["Specialising - Type"]
    rows = []
    for row in range(2, ws.max_row + 1):
        t = ws.cell(row, 1).value
        if not t or not str(t).strip():
            continue
        team_members = [ws.cell(row, c).value for c in range(5, 10)]
        team_members = [_mon_ref(m) for m in team_members if m]
        rows.append({
            "type": t,
            "mons_count": stats["type_mons_count"].get(t, 0),
            "used_count": stats["type_used_count"].get(t, 0),
            "ace": _mon_ref(ws.cell(row, 4).value),
            "team": team_members,
        })
    return rows


def export_specialising_gen(wb, stats):
    ws = wb["Specialising - Gen"]
    rows = []
    for row in range(2, ws.max_row + 1):
        g = ws.cell(row, 1).value
        if g is None or str(g).strip() == "":
            continue
        gen_key = str(g).strip()
        team_members = [ws.cell(row, c).value for c in range(5, 10)]
        team_members = [_mon_ref(m) for m in team_members if m]
        rows.append({
            "generation": gen_key,
            "mons_count": stats["gen_mons_count"].get(gen_key, 0),
            "used_count": stats["gen_used_count"].get(gen_key, 0),
            "ace": _mon_ref(ws.cell(row, 4).value),
            "team": team_members,
        })
    return rows


def export_gen_type_matrix(wb):
    ws = wb["Specialising - Gen_Type"]
    headers = [ws.cell(1, c).value for c in range(2, ws.max_column + 1) if ws.cell(1, c).value]
    rows = []
    for row in range(2, ws.max_row + 1):
        t = ws.cell(row, 1).value
        if not t or not str(t).strip():
            continue
        cells = {}
        for i, gen in enumerate(headers):
            val = ws.cell(row, 2 + i).value
            if val:
                cells[str(gen)] = {"name": val, "sprite": sprite_url(val)}
        rows.append({"type": t, "generations": cells})
    return {"generations": [str(g) for g in headers], "rows": rows}


def _read_move_catalog(ws):
    """Move name/type/split are plain typed-in values on the Moves sheet
    (columns A/B/C), unaffected by the formula-caching bug - read them
    directly."""
    catalog = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name or not str(name).strip():
            continue
        catalog.append({
            "move": str(name).strip(),
            "type": ws.cell(row, 2).value,
            "split": ws.cell(row, 3).value,
        })
    return catalog


def export_moves(wb, stats):
    ws = wb["Moves"]
    rows = []
    for m in _read_move_catalog(ws):
        rows.append({
            "move": m["move"],
            "type": m["type"],
            "split": m["split"],
            "count": stats["move_used_count"].get(m["move"], 0),
        })
    return rows


def export_moves_pivots(wb, stats):
    """Rebuilds the two former pivot tables (Sum of Count / Count of Moves,
    by Type & Split) directly from the Moves catalog (A/B/C, plain values)
    and the recomputed move-usage counts, instead of reading any Excel
    formula output - this sidesteps the caching bug entirely and needs no
    F:J / L:P summary columns on the sheet at all."""
    ws = wb["Moves"]
    catalog = _read_move_catalog(ws)

    types = sorted({m["type"] for m in catalog if m["type"]})
    splits = sorted({m["split"] for m in catalog if m["split"]})

    sum_of_count = Counter()   # (type, split) -> summed usage count
    count_of_moves = Counter()  # (type, split) -> number of moves

    for m in catalog:
        t, s = m["type"], m["split"]
        if not t or not s:
            continue
        count = stats["move_used_count"].get(m["move"], 0)
        sum_of_count[(t, s)] += count
        count_of_moves[(t, s)] += 1

    def build(counter):
        rows = []
        for t in types:
            entry = {"type": t}
            for s in splits:
                entry[s] = counter.get((t, s), 0)
            entry["total"] = sum(entry[s] for s in splits)
            rows.append(entry)
        return {"splits": splits, "rows": rows}

    return {
        "sumOfCount": build(sum_of_count),
        "countOfMoves": build(count_of_moves),
    }


def build_html(data):
    template_path = Path(TEMPLATE_FILE)
    if not template_path.exists():
        print(f"      Skipped HTML build: {TEMPLATE_FILE} not found next to this script.")
        return
    template = template_path.read_text(encoding="utf-8")
    html = template.replace("__DATA_JSON__", json.dumps(data))
    Path(HTML_OUTPUT_FILE).write_text(html, encoding="utf-8")
    print(f"      Wrote {HTML_OUTPUT_FILE} - open this in your browser")


def build_html_from_workbook():
    """Read the (now icon-updated) workbook and build pokemon_ui.html directly -
    no intermediate data.json file. The parsed data only ever exists as an
    in-memory dict; build_html() serializes it straight into the HTML."""
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    teams, summary = export_teams(wb)
    stats = build_roster_stats(teams)
    data = {
        "teams": teams,
        "summary": summary,
        "specialisingType": export_specialising_type(wb, stats),
        "specialisingGen": export_specialising_gen(wb, stats),
        "genTypeMatrix": export_gen_type_matrix(wb),
        "moves": export_moves(wb, stats),
        "movesPivots": export_moves_pivots(wb, stats),
    }
    total_mons = sum(len(t["pokemons"]) for t in teams)
    print(f"[1/2] Parsed {len(teams)} teams, {total_mons} pokemon rows, "
          f"{len(data['specialisingType'])} types, {len(data['specialisingGen'])} gens, "
          f"{len(data['moves'])} moves")

    build_html(data)


# ===========================================================================
# Entry point
# ===========================================================================

def main():
    # IMPORTANT: build the HTML first, while Pokemon.xlsx still has Excel's
    # cached formula results (Type/Gen Usage sheets, Moves pivot tables). If
    # add_icons_to_workbook() runs first, openpyxl re-saves the file and does
    # NOT preserve cached formula values - a later data_only=True read would
    # then see None for every formula-driven cell (usage counts, pivots).
    # Icons themselves don't depend on formula values, so they're safe to add
    # after - and Excel recalculates formulas on its own next time you open
    # the file, so this doesn't break anything in the workbook itself.
    build_html_from_workbook()
    add_icons_to_workbook()


if __name__ == "__main__":
    main()